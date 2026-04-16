#!/usr/bin/env python3
"""
Import RECURSOS and EQUIPOS from Excel to PostgreSQL
Handles category and unit matching from text to IDs
"""

import sys
import os
import warnings
from datetime import datetime

import openpyxl
import psycopg2
import psycopg2.extras

# Suppress openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# =============================================
# Configuration
# =============================================
DB_CONFIG = {
    "host": "localhost",
    "port": 5433,
    "user": "kardexchio",
    "password": "kardexchio_2026",
    "dbname": "kardexchio",
}

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "Kardex_Importacion_2026.xlsx"
)

# Unit mapping: Excel unit -> DB unit name (for fuzzy matching)
UNIT_MAPPING = {
    # Direct matches
    "und": "Unidad",
    "UND": "Unidad",
    "Und": "Unidad",
    "unidad": "Unidad",

    # Kilogram variants
    "kg": "Kilogramo",
    "Kg": "Kilogramo",
    "kilo": "Kilogramo",
    "Kilo": "Kilogramo",

    # Liter variants
    "l": "Litro",
    "litro": "Litro",
    "Litro": "Litro",

    # Meter variants
    "m": "Metro",
    "mt": "Metro",
    "Mt": "Metro",
    "metro": "Metro",
    "Metro": "Metro",

    # Meter squared
    "m2": "Metro cuadrado",
    "m²": "Metro cuadrado",

    # Gallon variants
    "gl": "Galón",
    "Gl": "Galón",
    "gal": "Galón",
    "Gal": "Galón",
    "galon": "Galón",
    "Galón": "Galón",

    # Pair variants
    "par": "Par",
    "Par": "Par",
    "pares": "Par",

    # Box
    "caja": "Caja",
    "Caja": "Caja",
    "cajas": "Caja",

    # Bag
    "bolsa": "Bolsa",
    "Bolsa": "Bolsa",
    "bolsas": "Bolsa",

    # Roll
    "rollo": "Rollo",
    "Rollo": "Rollo",
    "rollos": "Rollo",

    # Piece
    "pieza": "Pieza",
    "Pieza": "Pieza",
    "piezas": "Pieza",

    # Sheet
    "pliego": "Pliego",
    "Pliego": "Pliego",
    "pliegos": "Pliego",

    # Package/Packet variants (will default to Pieza if not found)
    "paquete": "Pieza",
    "Paquete": "Pieza",
    "pqt": "Pieza",
    "Pqt": "Pieza",
    "pk": "Pieza",
    "Pk": "Pieza",

    # Balde (bucket) - map to Pieza
    "balde": "Pieza",
    "Balde": "Pieza",
}


def safe_str(value, max_len=None):
    """Convert value to string safely, handling None and trimming."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if max_len:
        s = s[:max_len]
    return s


def load_category_map(cur):
    """Load categories from DB into a map (case-insensitive)."""
    cur.execute("SELECT id, nombre FROM categorias")
    cat_map = {}
    for row_id, nombre in cur.fetchall():
        # Store both exact case and lowercase for fuzzy matching
        cat_map[nombre] = row_id
        cat_map[nombre.lower()] = row_id
    return cat_map


def load_unit_map(cur):
    """Load units from DB into a map."""
    cur.execute("SELECT id, nombre FROM unidades_medida")
    unit_map = {}
    for row_id, nombre in cur.fetchall():
        # Store both exact case and lowercase
        unit_map[nombre] = row_id
        unit_map[nombre.lower()] = row_id
    return unit_map


def match_category(cat_text, cat_map):
    """Match category text to category ID from DB."""
    if not cat_text:
        return None

    cat_text = cat_text.strip()

    # Try exact match first
    if cat_text in cat_map:
        return cat_map[cat_text]

    # Try lowercase match
    if cat_text.lower() in cat_map:
        return cat_map[cat_text.lower()]

    return None


def match_unit(unit_text, unit_map):
    """Match unit text to unit ID from DB using mapping."""
    if not unit_text:
        return unit_map.get("Unidad")  # Default to Unidad

    unit_text = unit_text.strip()

    # Try UNIT_MAPPING first (fuzzy matching)
    if unit_text in UNIT_MAPPING:
        mapped_name = UNIT_MAPPING[unit_text]
        if mapped_name in unit_map:
            return unit_map[mapped_name]

    # Try exact match
    if unit_text in unit_map:
        return unit_map[unit_text]

    # Try lowercase match
    if unit_text.lower() in unit_map:
        return unit_map[unit_text.lower()]

    # Default to "Unidad" if not found
    return unit_map.get("Unidad")


def import_recursos(cur, cat_map, unit_map):
    """Import RECURSOS sheet."""
    print("\n=== Importing RECURSOS ===")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["RECURSOS"]

    recursos_data = []
    errors = []
    skipped = 0
    seen_codigos = set()

    # Start from row 4 (header is row 3)
    for i in range(4, ws.max_row + 1):
        codigo = safe_str(ws[i][0].value, 20)  # Column A
        nombre = safe_str(ws[i][1].value, 300)  # Column B
        categoria_text = safe_str(ws[i][2].value)  # Column C
        unidad_text = safe_str(ws[i][3].value)  # Column D

        # Skip if missing required fields
        if not codigo or not nombre:
            skipped += 1
            continue

        # Skip if duplicate codigo (keep first occurrence)
        if codigo in seen_codigos:
            errors.append(f"Row {i}: Duplicate codigo '{codigo}' (keeping first occurrence)")
            skipped += 1
            continue
        seen_codigos.add(codigo)

        # Match category
        categoria_id = match_category(categoria_text, cat_map)
        if not categoria_id:
            errors.append(f"Row {i}: Category '{categoria_text}' not found for '{nombre}'")
            skipped += 1
            continue

        # Match unit
        unidad_id = match_unit(unidad_text, unit_map)
        if not unidad_id:
            errors.append(f"Row {i}: Unit '{unidad_text}' not found for '{nombre}'")
            skipped += 1
            continue

        recursos_data.append((codigo, nombre, categoria_id, unidad_id))

    print(f"  Parsed {len(recursos_data)} recursos ({skipped} rows skipped)")

    if errors:
        print(f"  Errors found ({len(errors)} total, showing first 5):")
        for error in errors[:5]:
            print(f"    {error}")

    # Clear existing data (respecting foreign key constraints)
    cur.execute("DELETE FROM movimientos")
    cur.execute("DELETE FROM salidas")
    cur.execute("DELETE FROM entradas")
    cur.execute("DELETE FROM entrada_equipos")
    cur.execute("DELETE FROM salida_equipos")
    cur.execute("DELETE FROM equipos")
    cur.execute("DELETE FROM recursos")
    print("  Cleared existing data from dependent tables")

    # Reset sequences
    cur.execute("ALTER SEQUENCE recursos_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE entradas_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE salidas_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE equipos_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE salida_equipos_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE movimientos_id_seq RESTART WITH 1")

    # Batch insert
    if recursos_data:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO recursos (codigo, nombre, categoria_id, unidad_medida_id, activo)
               VALUES %s""",
            [(codigo, nombre, cat_id, unit_id, True) for codigo, nombre, cat_id, unit_id in recursos_data],
            template="(%s, %s, %s, %s, %s)",
        )
        print(f"  Inserted {len(recursos_data)} recursos")

    return len(recursos_data)


def import_equipos(cur, cat_map, unit_map):
    """Import EQUIPOS sheet."""
    print("\n=== Importing EQUIPOS ===")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["EQUIPOS"]

    equipos_data = []
    errors = []
    skipped = 0

    # Start from row 4 (header is row 3)
    for i in range(4, ws.max_row + 1):
        nombre = safe_str(ws[i][0].value, 300)  # Column A
        categoria_text = safe_str(ws[i][1].value)  # Column B
        unidad_text = safe_str(ws[i][2].value)  # Column C
        estado = safe_str(ws[i][3].value, 50)  # Column D (Estado Inicial)

        # Default estado if not provided
        if not estado:
            estado = "EN_ALMACEN"

        # Skip if missing nombre
        if not nombre:
            skipped += 1
            continue

        # Match category
        categoria_id = match_category(categoria_text, cat_map)
        if not categoria_id:
            errors.append(f"Row {i}: Category '{categoria_text}' not found for '{nombre}'")
            skipped += 1
            continue

        # Match unit (default to Unidad if not provided)
        unidad_id = match_unit(unidad_text, unit_map)
        if not unidad_id:
            unidad_id = unit_map.get("Unidad")

        equipos_data.append((nombre, categoria_id, unidad_id, estado))

    print(f"  Parsed {len(equipos_data)} equipos ({skipped} rows skipped)")

    if errors:
        print(f"  Errors found ({len(errors)} total, showing first 5):")
        for error in errors[:5]:
            print(f"    {error}")

    # Batch insert
    if equipos_data:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO equipos (nombre, categoria_id, unidad_medida_id, estado, activo)
               VALUES %s""",
            [(nombre, cat_id, unit_id, estado, True) for nombre, cat_id, unit_id, estado in equipos_data],
            template="(%s, %s, %s, %s, %s)",
        )
        print(f"  Inserted {len(equipos_data)} equipos")

    return len(equipos_data)


def verify_import(cur):
    """Verify and summarize imported data."""
    print("\n=== Verification ===")

    cur.execute("SELECT COUNT(*) FROM recursos")
    recursos_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM equipos")
    equipos_count = cur.fetchone()[0]

    print(f"  Recursos imported: {recursos_count}")
    print(f"  Equipos imported: {equipos_count}")

    if recursos_count > 0:
        print("\n  Sample recursos:")
        cur.execute("SELECT codigo, nombre FROM recursos LIMIT 5")
        for codigo, nombre in cur.fetchall():
            print(f"    {codigo}: {nombre[:50]}")

    if equipos_count > 0:
        print("\n  Sample equipos:")
        cur.execute("SELECT nombre, estado FROM equipos LIMIT 5")
        for nombre, estado in cur.fetchall():
            print(f"    {nombre[:50]}: {estado}")


def main():
    """Main entry point."""
    print("=" * 70)
    print("  KardexChio - Import RECURSOS and EQUIPOS from Excel")
    print("=" * 70)

    # Verify Excel file exists
    if not os.path.exists(EXCEL_PATH):
        print(f"\nERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"\nExcel file: {EXCEL_PATH}")

    # Connect to PostgreSQL
    print(f"Connecting to PostgreSQL at {DB_CONFIG['host']}:{DB_CONFIG['port']}...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.set_client_encoding("UTF8")
        print("  Connected successfully")
    except psycopg2.OperationalError as e:
        print(f"ERROR: Could not connect to PostgreSQL: {e}")
        sys.exit(1)

    try:
        cur = conn.cursor()

        # Load mappings
        print("\nLoading category and unit mappings from database...")
        cat_map = load_category_map(cur)
        unit_map = load_unit_map(cur)
        print(f"  Categories: {len([k for k in cat_map.keys() if k.isupper() or ' ' in k])} items")
        print(f"  Units: {len([k for k in unit_map.keys() if k.isupper() or ' ' in k])} items")

        # Import sheets
        recursos_count = import_recursos(cur, cat_map, unit_map)
        equipos_count = import_equipos(cur, cat_map, unit_map)

        # Commit
        conn.commit()
        print("\n[OK] All changes committed to database")

        # Verify
        verify_import(cur)

        print("\n" + "=" * 70)
        print(f"  IMPORT COMPLETED")
        print(f"  - Recursos: {recursos_count}")
        print(f"  - Equipos: {equipos_count}")
        print("=" * 70)

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: Import failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()
        print("\nDatabase connection closed.")


if __name__ == "__main__":
    main()
