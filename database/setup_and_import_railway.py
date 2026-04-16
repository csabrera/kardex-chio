#!/usr/bin/env python3
"""
Setup schema in Railway PostgreSQL and import data
"""

import sys
import os
import warnings
from urllib.parse import urlparse

import psycopg2
import psycopg2.extras
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# =============================================
# Configuration - Read from environment or use defaults
# =============================================
DATABASE_URL = os.environ.get("DATABASE_URL")

if DATABASE_URL:
    # Parse DATABASE_URL if available (from Railway)
    parsed = urlparse(DATABASE_URL)
    DB_CONFIG = {
        "host": parsed.hostname,
        "port": parsed.port or 5432,
        "user": parsed.username,
        "password": parsed.password,
        "dbname": parsed.path.lstrip('/'),
    }
else:
    # Fallback to hardcoded values
    DB_CONFIG = {
        "host": "postgres.railway.internal",
        "port": 5432,
        "user": "postgres",
        "password": "PaNcGXSppckjRACAflgpqdcZDkKNOkei",
        "dbname": "railway",
    }

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "Kardex_Importacion_2026.xlsx"
)

INIT_SQL_PATH = os.path.join(
    os.path.dirname(__file__),
    "init.sql"
)

UNIT_MAPPING = {
    "und": "Unidad", "UND": "Unidad", "Und": "Unidad", "unidad": "Unidad",
    "kg": "Kilogramo", "Kg": "Kilogramo", "kilo": "Kilogramo", "Kilo": "Kilogramo",
    "l": "Litro", "litro": "Litro", "Litro": "Litro",
    "m": "Metro", "mt": "Metro", "Mt": "Metro", "metro": "Metro", "Metro": "Metro",
    "m2": "Metro cuadrado", "m²": "Metro cuadrado",
    "gl": "Galón", "Gl": "Galón", "gal": "Galón", "Gal": "Galón",
    "par": "Par", "Par": "Par", "pares": "Par",
    "caja": "Caja", "Caja": "Caja", "cajas": "Caja",
    "bolsa": "Bolsa", "Bolsa": "Bolsa", "bolsas": "Bolsa",
    "rollo": "Rollo", "Rollo": "Rollo", "rollos": "Rollo",
    "pieza": "Pieza", "Pieza": "Pieza", "piezas": "Pieza",
    "pliego": "Pliego", "Pliego": "Pliego", "pliegos": "Pliego",
    "paquete": "Pieza", "Paquete": "Pieza", "pqt": "Pieza", "pk": "Pieza",
    "balde": "Pieza", "Balde": "Pieza",
}

def safe_str(value, max_len=None):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if max_len:
        s = s[:max_len]
    return s

def execute_schema(cur):
    """Execute init.sql to create schema"""
    print("\n=== Creating Schema ===")

    if not os.path.exists(INIT_SQL_PATH):
        print(f"ERROR: init.sql not found at {INIT_SQL_PATH}")
        return False

    with open(INIT_SQL_PATH, 'r', encoding='utf-8') as f:
        sql_content = f.read()

    # Split by semicolon and execute each statement
    statements = sql_content.split(';')
    count = 0

    for statement in statements:
        statement = statement.strip()
        if not statement:
            continue

        try:
            cur.execute(statement)
            count += 1
        except psycopg2.errors.DuplicateTable as e:
            # Table already exists - skip
            pass
        except Exception as e:
            print(f"  Warning executing statement: {e}")
            continue

    print(f"  [OK] Schema created ({count} statements executed)")
    return True

def load_category_map(cur):
    cur.execute("SELECT id, nombre FROM categorias")
    cat_map = {}
    for row_id, nombre in cur.fetchall():
        cat_map[nombre] = row_id
        cat_map[nombre.lower()] = row_id
    return cat_map

def load_unit_map(cur):
    cur.execute("SELECT id, nombre FROM unidades_medida")
    unit_map = {}
    for row_id, nombre in cur.fetchall():
        unit_map[nombre] = row_id
        unit_map[nombre.lower()] = row_id
    return unit_map

def match_category(cat_text, cat_map):
    if not cat_text:
        return None
    cat_text = cat_text.strip()
    if cat_text in cat_map:
        return cat_map[cat_text]
    if cat_text.lower() in cat_map:
        return cat_map[cat_text.lower()]
    return None

def match_unit(unit_text, unit_map):
    if not unit_text:
        return unit_map.get("Unidad")
    unit_text = unit_text.strip()
    if unit_text in UNIT_MAPPING:
        mapped_name = UNIT_MAPPING[unit_text]
        if mapped_name in unit_map:
            return unit_map[mapped_name]
    if unit_text in unit_map:
        return unit_map[unit_text]
    if unit_text.lower() in unit_map:
        return unit_map[unit_text.lower()]
    return unit_map.get("Unidad")

def import_recursos(cur, cat_map, unit_map):
    print("\n=== Importing RECURSOS ===")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["RECURSOS"]

    recursos_data = []
    errors = []
    skipped = 0
    seen_codigos = set()

    for i in range(4, ws.max_row + 1):
        codigo = safe_str(ws[i][0].value, 20)
        nombre = safe_str(ws[i][1].value, 300)
        categoria_text = safe_str(ws[i][2].value)
        unidad_text = safe_str(ws[i][3].value)

        if not codigo or not nombre:
            skipped += 1
            continue

        if codigo in seen_codigos:
            errors.append(f"Duplicate codigo '{codigo}'")
            skipped += 1
            continue
        seen_codigos.add(codigo)

        categoria_id = match_category(categoria_text, cat_map)
        if not categoria_id:
            errors.append(f"Category '{categoria_text}' not found")
            skipped += 1
            continue

        unidad_id = match_unit(unidad_text, unit_map)
        if not unidad_id:
            unidad_id = unit_map.get("Unidad")

        recursos_data.append((codigo, nombre, categoria_id, unidad_id))

    print(f"  Parsed {len(recursos_data)} recursos ({skipped} rows skipped)")

    if recursos_data:
        psycopg2.extras.execute_values(
            cur,
            "INSERT INTO recursos (codigo, nombre, categoria_id, unidad_medida_id, activo) VALUES %s",
            [(codigo, nombre, cat_id, unit_id, True) for codigo, nombre, cat_id, unit_id in recursos_data],
            template="(%s, %s, %s, %s, %s)",
        )
        print(f"  [OK] Inserted {len(recursos_data)} recursos")

    return len(recursos_data)

def import_equipos(cur, cat_map, unit_map):
    print("\n=== Importing EQUIPOS ===")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["EQUIPOS"]

    equipos_data = []
    errors = []
    skipped = 0

    for i in range(4, ws.max_row + 1):
        nombre = safe_str(ws[i][0].value, 300)
        categoria_text = safe_str(ws[i][1].value)
        unidad_text = safe_str(ws[i][2].value)
        estado = safe_str(ws[i][3].value, 50) or "EN_ALMACEN"

        if not nombre:
            skipped += 1
            continue

        categoria_id = match_category(categoria_text, cat_map)
        if not categoria_id:
            errors.append(f"Category '{categoria_text}' not found")
            skipped += 1
            continue

        unidad_id = match_unit(unidad_text, unit_map)
        if not unidad_id:
            unidad_id = unit_map.get("Unidad")

        equipos_data.append((nombre, categoria_id, unidad_id, estado))

    print(f"  Parsed {len(equipos_data)} equipos ({skipped} rows skipped)")

    if equipos_data:
        psycopg2.extras.execute_values(
            cur,
            "INSERT INTO equipos (nombre, categoria_id, unidad_medida_id, estado, activo) VALUES %s",
            [(nombre, cat_id, unit_id, estado, True) for nombre, cat_id, unit_id, estado in equipos_data],
            template="(%s, %s, %s, %s, %s)",
        )
        print(f"  [OK] Inserted {len(equipos_data)} equipos")

    return len(equipos_data)

def verify_import(cur):
    print("\n=== Verification ===")
    cur.execute("SELECT COUNT(*) FROM recursos")
    recursos_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM equipos")
    equipos_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM categorias")
    cat_count = cur.fetchone()[0]

    print(f"  Categorias: {cat_count}")
    print(f"  Recursos: {recursos_count}")
    print(f"  Equipos: {equipos_count}")

def main():
    print("=" * 70)
    print("  KardexChio - Setup Schema + Import to Railway")
    print("=" * 70)
    print(f"\nConnecting to Railway PostgreSQL...")
    print(f"  Host: {DB_CONFIG['host']}:{DB_CONFIG['port']}")
    print(f"  Database: {DB_CONFIG['dbname']}")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.set_client_encoding("UTF8")
        print("  [OK] Connected")
    except Exception as e:
        print(f"\n[ERROR] Could not connect to Railway:")
        print(f"  {e}")
        print("\nMake sure:")
        print("  1. Railway PostgreSQL is running")
        print("  2. DATABASE_URL is correct")
        print("  3. You have internet connection")
        sys.exit(1)

    try:
        cur = conn.cursor()

        # Step 1: Create schema
        execute_schema(cur)
        conn.commit()

        # Step 2: Load mappings
        print("\n=== Loading Mappings ===")
        cat_map = load_category_map(cur)
        unit_map = load_unit_map(cur)
        print(f"  Categories: {len([k for k in cat_map.keys() if ' ' in str(k) or k.isupper()])} items")
        print(f"  Units: {len([k for k in unit_map.keys() if ' ' in str(k) or k.isupper()])} items")

        # Step 3: Import data
        recursos_count = import_recursos(cur, cat_map, unit_map)
        equipos_count = import_equipos(cur, cat_map, unit_map)

        # Commit all changes
        conn.commit()
        print("\n[OK] All changes committed to Railway PostgreSQL")

        # Verify
        verify_import(cur)

        print("\n" + "=" * 70)
        print("  SUCCESS - Setup and Import Complete!")
        print(f"  - Schema created")
        print(f"  - Recursos: {recursos_count}")
        print(f"  - Equipos: {equipos_count}")
        print("=" * 70)
        print("\nYour Railway PostgreSQL is now ready!")
        print("Backend should now be able to start correctly.")

    except Exception as e:
        conn.rollback()
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()

if __name__ == "__main__":
    main()
