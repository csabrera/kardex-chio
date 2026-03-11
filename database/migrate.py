#!/usr/bin/env python3
"""
KardexChio - Data Migration Script
Reads Excel file and imports all data into PostgreSQL.
"""

import sys
import os
import warnings
from datetime import datetime, date

import openpyxl
import psycopg2
import psycopg2.extras

# Suppress openpyxl warnings about unsupported extensions
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# =============================================
# Configuration
# =============================================
DB_CONFIG = {
    "host": "localhost",
    "port": 5433,
    "user": "kardexchio",
    "password": "kardexchio_2026",
    "dbname": "kardexchio",
}

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Kardex fica 2026.xlsx")

# Category prefixes for generating codes when missing
CATEGORY_PREFIXES = {
    "Ferreteria": "Fer",
    "Pintura": "Pin",
    "Electrico": "Ele",
    "Seguridad": "Seg",
    "Maderas": "Mad",
    "Herramientas": "Her",
    "Limpieza": "Lim",
    "Ladrillos": "Lad",
    "Combustibles": "Com",
    "Agua": "Agu",
    "Gaseosas": "Gas",
    "Galletas": "Gal",
    "Microcemento": "Mic",
    "Oficina": "Ofi",
    "Equipo": "Equ",
}

EXPECTED_CATEGORIES = [
    "Ferreteria", "Pintura", "Electrico", "Seguridad", "Maderas",
    "Herramientas", "Limpieza", "Ladrillos", "Combustibles", "Agua",
    "Gaseosas", "Galletas", "Microcemento", "Oficina", "Equipo",
]


def safe_str(value, max_len=None):
    """Convert value to string safely, handling None and trimming."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if max_len:
        s = s[:max_len]
    return s


def safe_int(value, default=0):
    """Convert value to int safely."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def safe_decimal(value, default=0):
    """Convert value to float safely."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def to_date(value):
    """Convert datetime/date value to date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Try parsing string
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def read_excel(excel_path):
    """Read the Excel file and return the workbook."""
    print(f"Reading Excel file: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"  Sheets found: {wb.sheetnames}")
    return wb


def step1_verify_categories(cur, wb):
    """Step 1: Verify categories exist, add missing ones."""
    print("\n=== Step 1: Verifying categories ===")

    # Get existing categories from DB
    cur.execute("SELECT id, nombre FROM categorias ORDER BY id")
    existing = {row[1]: row[0] for row in cur.fetchall()}
    print(f"  Existing categories in DB: {len(existing)}")

    # Collect categories from Excel INVENTARIO sheet
    ws = wb["INVENTARIO"]
    excel_cats = set()
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=False):
        cat = row[5].value  # Column F = CATEGORÍA (0-indexed: col 6, but iter_rows gives 0-based)
        if cat and str(cat).strip():
            excel_cats.add(str(cat).strip())

    # Insert missing categories
    missing = excel_cats - set(existing.keys())
    if missing:
        print(f"  Adding {len(missing)} missing categories: {missing}")
        for cat_name in sorted(missing):
            cur.execute(
                "INSERT INTO categorias (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING RETURNING id",
                (cat_name,)
            )
    else:
        print("  All categories already exist.")

    # Refresh category map
    cur.execute("SELECT id, nombre FROM categorias ORDER BY id")
    cat_map = {row[1]: row[0] for row in cur.fetchall()}
    print(f"  Total categories: {len(cat_map)}")
    return cat_map


def step2_import_recursos(cur, wb, cat_map):
    """Step 2: Import INVENTARIO sheet into recursos table."""
    print("\n=== Step 2: Importing recursos (INVENTARIO) ===")

    ws = wb["INVENTARIO"]
    # Headers at row 10 (0-indexed cols):
    #   B=Nº(1), C=NUM(2), D=CÓDIGO(3), E=RECURSO(4), F=CATEGORÍA(5),
    #   G=UNIDAD(6), H=STATUS(7), I=ALERTA(8), J=STOCK_MINIMO(9)

    recursos_data = []
    seen_codigos = set()
    seen_nombres = set()
    skipped = 0

    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=False):
        recurso_name = safe_str(row[4].value)  # col E = RECURSO
        if not recurso_name:
            skipped += 1
            continue

        # Skip duplicate nombres
        if recurso_name in seen_nombres:
            skipped += 1
            continue
        seen_nombres.add(recurso_name)

        num = safe_int(row[2].value, 0)  # col C = NUM
        codigo = safe_str(row[3].value, 20)  # col D = CÓDIGO
        categoria = safe_str(row[5].value)  # col F = CATEGORÍA
        unidad = safe_str(row[6].value, 50) or "Und"  # col G = UNIDAD
        stock_minimo = safe_int(row[9].value, 0)  # col J = STOCK MÍNIMO

        # Look up categoria_id
        cat_id = cat_map.get(categoria)
        if not cat_id:
            print(f"  WARNING: Category '{categoria}' not found for recurso '{recurso_name}', skipping")
            skipped += 1
            continue

        # Generate codigo if missing
        if not codigo:
            prefix = CATEGORY_PREFIXES.get(categoria, "XXX")
            codigo = f"{prefix}{num:04d}"

        # Ensure unique codigo
        if codigo in seen_codigos:
            # Append a suffix to make it unique
            suffix = 1
            while f"{codigo}_{suffix}" in seen_codigos:
                suffix += 1
            codigo = f"{codigo}_{suffix}"
        seen_codigos.add(codigo)

        recursos_data.append((num, codigo, recurso_name, cat_id, unidad, stock_minimo))

    print(f"  Parsed {len(recursos_data)} recursos ({skipped} rows skipped)")

    # Clear existing data (idempotent)
    cur.execute("DELETE FROM movimientos")
    cur.execute("DELETE FROM salidas")
    cur.execute("DELETE FROM entradas")
    cur.execute("DELETE FROM salida_equipos")
    cur.execute("DELETE FROM equipos")
    cur.execute("DELETE FROM recursos")
    print("  Cleared existing data from dependent tables")

    # Reset sequences
    cur.execute("ALTER SEQUENCE recursos_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE entradas_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE salidas_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE equipos_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE salida_equipos_id_seq RESTART WITH 1")
    cur.execute("ALTER SEQUENCE movimientos_id_seq RESTART WITH 1")

    # Batch insert
    BATCH_SIZE = 100
    inserted = 0
    for i in range(0, len(recursos_data), BATCH_SIZE):
        batch = recursos_data[i:i + BATCH_SIZE]
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO recursos (num, codigo, nombre, categoria_id, unidad, stock_minimo)
               VALUES %s
               ON CONFLICT (codigo) DO NOTHING""",
            batch,
            template="(%s, %s, %s, %s, %s, %s)",
        )
        inserted += len(batch)
        if inserted % 500 == 0 or inserted == len(recursos_data):
            print(f"  Importing recursos... {inserted}/{len(recursos_data)}")

    # Build recurso lookup map by nombre
    cur.execute("SELECT id, nombre FROM recursos")
    recurso_map = {row[1]: row[0] for row in cur.fetchall()}
    print(f"  Inserted {len(recurso_map)} recursos into database")
    return recurso_map


def step3_import_entradas(cur, wb, recurso_map):
    """Step 3: Import ENTRADAS sheet."""
    print("\n=== Step 3: Importing entradas ===")

    ws = wb["ENTRADAS"]
    # Headers at row 8 (0-indexed cols):
    #   B=Nº(1), C=FECHA(2), D=Nº GUÍA(3), E=CÓDIGO(4), F=RECURSO(5),
    #   G=CATEGORÍA(6), H=UNIDAD(7), I=EXISTENCIA(8), J=CANTIDAD(9),
    #   K=QUIEN_ENTREGA(10), L=QUIEN_RECIBE(11), M=MEDIO_TRANSPORTE(12)

    entradas_data = []
    skipped = 0
    warnings_list = []

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=False):
        fecha_val = row[2].value  # col C = FECHA
        recurso_name = safe_str(row[5].value)  # col F = RECURSO

        if not fecha_val or not recurso_name:
            continue

        fecha = to_date(fecha_val)
        if not fecha:
            skipped += 1
            continue

        cantidad = safe_decimal(row[9].value, 0)  # col J = CANTIDAD
        if cantidad <= 0:
            skipped += 1
            continue

        recurso_id = recurso_map.get(recurso_name)
        if not recurso_id:
            if recurso_name not in [w[0] for w in warnings_list]:
                warnings_list.append((recurso_name, row[0].row))
            skipped += 1
            continue

        num_guia = safe_str(row[3].value, 50)  # col D = Nº GUÍA
        quien_entrega = safe_str(row[10].value, 150)  # col K
        quien_recibe = safe_str(row[11].value, 150)  # col L
        medio_transporte = safe_str(row[12].value, 150)  # col M

        entradas_data.append((fecha, num_guia, recurso_id, cantidad, quien_entrega, quien_recibe, medio_transporte))

    if warnings_list:
        print(f"  WARNING: {len(warnings_list)} recursos not found in DB:")
        for name, row_num in warnings_list[:5]:
            print(f"    - '{name}' (first seen at row {row_num})")

    print(f"  Parsed {len(entradas_data)} entradas ({skipped} rows skipped)")

    # Batch insert
    BATCH_SIZE = 200
    inserted = 0
    for i in range(0, len(entradas_data), BATCH_SIZE):
        batch = entradas_data[i:i + BATCH_SIZE]
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO entradas (fecha, num_guia, recurso_id, cantidad, quien_entrega, quien_recibe, medio_transporte)
               VALUES %s""",
            batch,
            template="(%s, %s, %s, %s, %s, %s, %s)",
        )
        inserted += len(batch)
        if inserted % 500 == 0 or inserted == len(entradas_data):
            print(f"  Importing entradas... {inserted}/{len(entradas_data)}")

    print(f"  Inserted {len(entradas_data)} entradas")
    return len(entradas_data)


def step4_import_salidas(cur, wb, recurso_map):
    """Step 4: Import SALIDAS sheet."""
    print("\n=== Step 4: Importing salidas ===")

    ws = wb["SALIDAS"]
    # Headers at row 8 (0-indexed cols):
    #   B=Nº(1), C=FECHA(2), D=Nº REGISTRO(3), E=CÓDIGO(4), F=RECURSO(5),
    #   G=CATEGORÍA(6), H=UNIDAD(7), I=EXISTENCIA(8), J=CANTIDAD(9),
    #   K=FRENTE_TRABAJO(10), L=DESCRIPCION(11), M=QUIEN_ENTREGA(12), N=QUIEN_RECIBE(13)

    salidas_data = []
    skipped = 0
    warnings_list = []

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=False):
        fecha_val = row[2].value  # col C = FECHA
        recurso_name = safe_str(row[5].value)  # col F = RECURSO

        if not fecha_val or not recurso_name:
            continue

        fecha = to_date(fecha_val)
        if not fecha:
            skipped += 1
            continue

        cantidad = safe_decimal(row[9].value, 0)  # col J = CANTIDAD
        if cantidad <= 0:
            skipped += 1
            continue

        recurso_id = recurso_map.get(recurso_name)
        if not recurso_id:
            if recurso_name not in [w[0] for w in warnings_list]:
                warnings_list.append((recurso_name, row[0].row))
            skipped += 1
            continue

        # Nº REGISTRO column (D) - may contain CÓDIGO or be empty
        num_registro = safe_str(row[3].value, 50)  # col D
        frente_trabajo = safe_str(row[10].value, 150)  # col K
        descripcion = safe_str(row[11].value, 300)  # col L
        quien_entrega = safe_str(row[12].value, 150)  # col M
        quien_recibe = safe_str(row[13].value, 150)  # col N

        salidas_data.append((fecha, num_registro, recurso_id, cantidad, frente_trabajo, descripcion, quien_entrega, quien_recibe))

    if warnings_list:
        print(f"  WARNING: {len(warnings_list)} recursos not found in DB:")
        for name, row_num in warnings_list[:5]:
            print(f"    - '{name}' (first seen at row {row_num})")

    print(f"  Parsed {len(salidas_data)} salidas ({skipped} rows skipped)")

    # Batch insert
    BATCH_SIZE = 200
    inserted = 0
    for i in range(0, len(salidas_data), BATCH_SIZE):
        batch = salidas_data[i:i + BATCH_SIZE]
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO salidas (fecha, num_registro, recurso_id, cantidad, frente_trabajo, descripcion_trabajo, quien_entrega, quien_recibe)
               VALUES %s""",
            batch,
            template="(%s, %s, %s, %s, %s, %s, %s, %s)",
        )
        inserted += len(batch)
        if inserted % 500 == 0 or inserted == len(salidas_data):
            print(f"  Importing salidas... {inserted}/{len(salidas_data)}")

    print(f"  Inserted {len(salidas_data)} salidas")
    return len(salidas_data)


def step5_import_equipos(cur, wb, cat_map):
    """Step 5: Import Hoja1 (equipos table)."""
    print("\n=== Step 5: Importing equipos (Hoja1) ===")

    ws = wb["Hoja1"]
    # Headers at row 2: ITEM(A/1), col B = equipment name, EN ALMACEN(C/3), SALIDA(D/4), INGRESO(E/5)
    # Data starts at row 3, equipment names are in column B

    equipo_cat_id = cat_map.get("Equipo")
    if not equipo_cat_id:
        print("  WARNING: 'Equipo' category not found, skipping equipos import")
        return

    equipos_data = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        nombre = safe_str(row[1].value, 300) if len(row) > 1 else None  # col B
        if not nombre:
            continue
        equipos_data.append((nombre, equipo_cat_id, "Und", "EN_ALMACEN"))

    print(f"  Parsed {len(equipos_data)} equipos")

    if equipos_data:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO equipos (nombre, categoria_id, unidad, estado)
               VALUES %s""",
            equipos_data,
            template="(%s, %s, %s, %s)",
        )
        print(f"  Inserted {len(equipos_data)} equipos")
    else:
        print("  No equipos to insert")


def step6_verify(cur):
    """Step 6: Verification - print summary counts and spot checks."""
    print("\n=== Step 6: Verification ===")

    cur.execute("SELECT COUNT(*) FROM recursos")
    recursos_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM entradas")
    entradas_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM salidas")
    salidas_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM equipos")
    equipos_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM categorias")
    categorias_count = cur.fetchone()[0]

    print(f"\n  {'Table':<20} {'Count':>8}  {'Expected':>10}")
    print(f"  {'-'*42}")
    print(f"  {'categorias':<20} {categorias_count:>8}  {'~15':>10}")
    print(f"  {'recursos':<20} {recursos_count:>8}  {'~1196':>10}")
    print(f"  {'entradas':<20} {entradas_count:>8}  {'~2906+':>10}")
    print(f"  {'salidas':<20} {salidas_count:>8}  {'~1734':>10}")
    print(f"  {'equipos':<20} {equipos_count:>8}  {'~2':>10}")

    # Spot-check: top 5 recursos by existencia_actual
    print("\n  Top 5 recursos by existencia_actual (from vista_inventario):")
    cur.execute("""
        SELECT codigo, nombre, total_entradas, total_salidas, existencia_actual, status
        FROM vista_inventario
        ORDER BY existencia_actual DESC
        LIMIT 5
    """)
    for row in cur.fetchall():
        print(f"    {row[0]}: {row[1][:50]:<50} E={row[2]:>6} S={row[3]:>6} Stock={row[4]:>6} [{row[5]}]")

    # Check recursos with AGOTADO status
    cur.execute("SELECT COUNT(*) FROM vista_inventario WHERE status = 'AGOTADO'")
    agotado = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vista_inventario WHERE status = 'MENOR AL MINIMO'")
    menor = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vista_inventario WHERE status = 'OK'")
    ok = cur.fetchone()[0]
    print(f"\n  Status distribution: OK={ok}, MENOR AL MINIMO={menor}, AGOTADO={agotado}")

    print("\n  Migration completed successfully!")


def main():
    """Main entry point."""
    print("=" * 60)
    print("  KardexChio - Data Migration")
    print("=" * 60)

    # Read Excel
    wb = read_excel(EXCEL_PATH)

    # Connect to PostgreSQL
    print(f"\nConnecting to PostgreSQL at {DB_CONFIG['host']}:{DB_CONFIG['port']}...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.set_client_encoding("UTF8")
        print("  Connected successfully")
    except psycopg2.OperationalError as e:
        print(f"ERROR: Could not connect to PostgreSQL: {e}")
        print("Make sure Docker is running and the database container is up.")
        sys.exit(1)

    try:
        cur = conn.cursor()

        # Step 1: Verify categories
        cat_map = step1_verify_categories(cur, wb)

        # Step 2: Import recursos
        recurso_map = step2_import_recursos(cur, wb, cat_map)

        # Step 3: Import entradas
        step3_import_entradas(cur, wb, recurso_map)

        # Step 4: Import salidas
        step4_import_salidas(cur, wb, recurso_map)

        # Step 5: Import equipos
        step5_import_equipos(cur, wb, cat_map)

        # Commit all changes
        conn.commit()
        print("\n  All changes committed to database.")

        # Step 6: Verification (read-only)
        step6_verify(cur)

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: Migration failed: {e}")
        print("  All changes have been rolled back.")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()
        print("\nDatabase connection closed.")


if __name__ == "__main__":
    main()
