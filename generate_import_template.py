#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# Mapeo de categorías Excel -> Sistema
mapeo_categorias = {
    'Ferreteria': 'Ferretería',
    'Electrico': 'Eléctrico',
    'Equipo': 'Equipos',
    'Herramientas': 'Herramientas',
    'Seguridad': 'Seguridad',
    'Maderas': 'Maderas',
    'Limpieza': 'Limpieza',
    'Oficina': 'Oficina',
    'Pintura': 'Pintura',
    'Combustibles': 'Combustibles',
    'Agua': 'Alimentos y Bebidas',
    'Galletas': 'Alimentos y Bebidas',
    'Gaseosas': 'Alimentos y Bebidas',
    'Ladrillos': 'Materiales de Construcción',
    'Microcemento': 'Materiales de Construcción'
}

# Categorías que son RECURSOS (se gastan)
recursos_cats = ['Ferreteria', 'Pintura', 'Electrico', 'Combustibles', 'Limpieza',
                 'Maderas', 'Oficina', 'Agua', 'Galletas', 'Gaseosas', 'Ladrillos', 'Microcemento']

# Categorías que son EQUIPOS (se prestan/devuelven)
equipos_cats = ['Herramientas', 'Seguridad', 'Equipo']

# Crear un ExcelWriter
with pd.ExcelWriter('Kardex_Importacion_Template_2026.xlsx', engine='openpyxl') as writer:

    # Leer datos del Kardex
    df_inv = pd.read_excel('Kardex fica 2026.xlsx', sheet_name='INVENTARIO', skiprows=9)

    # ====== HOJA 1: RECURSOS (solo items que se gastan) ======
    df_inv_recursos = df_inv[df_inv['CATEGORÍA'].isin(recursos_cats)].copy()

    df_recursos = df_inv_recursos[['CÓDIGO', 'RECURSO', 'CATEGORÍA', 'UNIDAD', 'EXISTENCIA ACTUAL']].copy()
    df_recursos.columns = ['Código', 'Nombre', 'Categoría', 'Unidad', 'Stock Inicial']
    df_recursos['Categoría'] = df_recursos['Categoría'].map(mapeo_categorias)
    df_recursos = df_recursos.dropna(subset=['Código', 'Nombre'])
    df_recursos = df_recursos[df_recursos['Código'].astype(str).str.strip() != '']
    df_recursos = df_recursos.reset_index(drop=True)

    df_recursos.to_excel(writer, sheet_name='RECURSOS', index=False, startrow=2)

    # Formatear hoja RECURSOS
    ws = writer.sheets['RECURSOS']
    ws['A1'] = 'IMPORTACION DE RECURSOS (se gastan)'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:E1')

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # ====== HOJA 2: ENTRADAS_RECURSOS ======
    df_ent_raw = pd.read_excel('Kardex fica 2026.xlsx', sheet_name='ENTRADAS', skiprows=7)
    df_ent_raw = df_ent_raw.dropna(how='all')

    # Seleccionar por índice (para evitar problemas con encoding)
    df_entradas = df_ent_raw.head(3000).iloc[:, [2, 3, 4, 5, 9, 10, 11, 12]].copy()
    df_entradas.columns = ['Fecha', 'Nº Guía', 'Código', 'Recurso', 'Cantidad',
                           'Quién Entrega (Nombre)', 'Quién Recibe (Nombre)', 'Medio Transporte']
    df_entradas = df_entradas[df_entradas['Código'].notna()].reset_index(drop=True)
    df_entradas = df_entradas[df_entradas['Código'].astype(str).str.strip() != ''].reset_index(drop=True)

    df_entradas.to_excel(writer, sheet_name='ENTRADAS_RECURSOS', index=False, startrow=2)

    ws = writer.sheets['ENTRADAS_RECURSOS']
    ws['A1'] = 'IMPORTACION DE ENTRADAS DE RECURSOS'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:H1')

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20

    # ====== HOJA 3: EQUIPOS (solo items que se prestan) ======
    df_inv_equipos = df_inv[df_inv['CATEGORÍA'].isin(equipos_cats)].copy()

    df_equipos_final = df_inv_equipos[['CÓDIGO', 'RECURSO', 'CATEGORÍA', 'UNIDAD', 'EXISTENCIA ACTUAL']].copy()
    df_equipos_final.columns = ['Código', 'Nombre', 'Categoría', 'Unidad', 'Cantidad Inicial']
    df_equipos_final['Categoría'] = df_equipos_final['Categoría'].map(mapeo_categorias)
    df_equipos_final['Estado Inicial'] = 'EN_ALMACEN'
    df_equipos_final = df_equipos_final[['Código', 'Nombre', 'Categoría', 'Unidad', 'Cantidad Inicial', 'Estado Inicial']]
    df_equipos_final = df_equipos_final.dropna(subset=['Código', 'Nombre'])
    df_equipos_final = df_equipos_final[df_equipos_final['Código'].astype(str).str.strip() != '']
    df_equipos_final = df_equipos_final.reset_index(drop=True)

    df_equipos_final.to_excel(writer, sheet_name='EQUIPOS', index=False, startrow=2)

    ws = writer.sheets['EQUIPOS']
    ws['A1'] = 'IMPORTACION DE EQUIPOS (se prestan)'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 15

    # ====== HOJA 4: SALIDA_EQUIPOS ======
    df_sal_raw = pd.read_excel('Kardex fica 2026.xlsx', sheet_name='SALIDA EQUIPOS', skiprows=7)
    df_sal_raw = df_sal_raw.dropna(how='all')

    # Tomar primeras 2000 filas usando índices de columna
    df_salida = df_sal_raw.head(2000).iloc[:, [2, 4, 5, 9, 10, 11, 12, 13]].copy()
    df_salida.columns = ['Fecha', 'Código Equipo', 'Nombre Equipo', 'Cantidad',
                         'Frente Trabajo', 'Descripción', 'Quién Entrega (Nombre)', 'Quién Recibe (Nombre)']
    df_salida = df_salida[df_salida['Código Equipo'].notna()].reset_index(drop=True)
    df_salida = df_salida[df_salida['Código Equipo'].astype(str).str.strip() != ''].reset_index(drop=True)

    # Insertar columna Tipo
    df_salida.insert(4, 'Tipo (SALIDA/INGRESO)', 'SALIDA')

    df_salida.to_excel(writer, sheet_name='SALIDA_EQUIPOS', index=False, startrow=2)

    ws = writer.sheets['SALIDA_EQUIPOS']
    ws['A1'] = 'IMPORTACION DE SALIDAS/ENTRADAS DE EQUIPOS'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:I1')

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

print("[OK] Excel template creado: Kardex_Importacion_Template_2026.xlsx")
print()
print("RESUMEN FINAL:")
print(f"  RECURSOS (se gastan):     {len(df_recursos)} items")
print(f"  ENTRADAS_RECURSOS:        {len(df_entradas)} movimientos")
print(f"  EQUIPOS (se prestan):     {len(df_equipos_final)} items")
print(f"  SALIDA_EQUIPOS:           {len(df_salida)} movimientos")
print()
print("Ubicacion: Kardex_Importacion_Template_2026.xlsx")
